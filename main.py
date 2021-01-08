# Process:
# Import necessary libraries
# Get spreadsheet location and name
# Store values in variables
# Make the input text of the name of the stock uppercase 
# Present a small menu which will allow to either edit, add, or view
# Create functions for each with proper logic

import openpyxl
import os

def setup():
    directory = "C:\\Users\\shrey\\Desktop"
    directory = directory.lower()
    os.chdir(directory)
    spreadname = "Stocks.xlsx"
    workbook = openpyxl.load_workbook(spreadname)
    sheet_Arr = workbook.sheetnames
    answer = ''
    sheetname = ''
    while answer != 'Y':
        for x in sheet_Arr:
            answer = input(x + ": Is this the sheetname which you are trying to access? (Y/N)  ")
            answer = answer.upper()
            if (answer == 'Y'):
                sheetname = sheet_Arr[0]
                break
    print(sheetname)
    sheet = workbook[sheetname]
    choose(sheet, workbook)

def choose(sheet, workbook):
    choice = ''
    while choice != 'V' or choice != 'E' or choice != 'A': 
        choice = input("Would you like to view (V), edit (E), or add (A) a stock?  ")
        choice = choice.upper()
        if choice == 'V':
            showStock(sheet)
            break
        elif choice == 'E':
            editStock(sheet, workbook)
            break
        elif choice == 'A':
            addStock(sheet, workbook)
            break
        else:
            print("Please enter a valid option") 
    

def showStock(sheet):
    stockName = input("What is the name of the stock which you are trying to view/edit? Example: TSLA    Stock Symbol: ")
    stockName = stockName.upper()
    stockRow = 0
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column = 2).value == stockName:
            stockRow = i
    print('\n')
    print("Here is the info for " + str(stockName) + ":")
    for i in range(2, sheet.max_column + 1):
        if i == 8 or i == 3 or i == 2 or i == 1:
            print(str((i - 1)) + ": " + sheet.cell(row = 1, column = i).value + ": " + str(sheet.cell(row = stockRow, column = i).value))
        else:
             print(str((i - 1)) + ": " + sheet.cell(row = 1, column = i).value + ": " + str(float((sheet.cell(row = stockRow, column = i).value))))
    return stockRow


def editStock(sheet, workbook):
    print("Here is a current list of your stocks: ")
    for i in range(2, sheet.max_row + 1):
        print("Investment Name: " + str(sheet.cell(row = i, column = 1).value) + '\n' + "Stock Symbol: " + str(sheet.cell(row = i, column = 2).value))
        print('\n')
    stockRow = showStock(sheet)
    choice = int(input("Which line would you like to edit? "))
    print(choice)
    if choice == 7 or choice == 2 or choice == 1:
        print("Current: " + str(sheet.cell(row = stockRow, column = choice + 1).value))
        edit = input("what would you like to change it to? ")
        sheet.cell(row = stockRow, column = choice + 1).value = edit
    else:
        print("Current: " + str(float(sheet.cell(row = stockRow, column = choice + 1).value)))
        edit = input("what would you like to change it to? ")
        sheet.cell(row = stockRow, column = choice + 1).value = float(edit)
    directory = "C:\\Users\\shrey\\Desktop"
    directory = directory.lower()
    os.chdir(directory)
    workbook.save("Stocks.xlsx")

def addStock(sheet, workbook):
    line = sheet.max_row + 1
    for i in range(1, sheet.max_column + 1):
        if i == 7 or i == 2 or i == 1:
            add = input(sheet.cell(row = 1, column = i).value + ": ")
            sheet.cell(row = line, column = i).value = add
        else:
            add = input(sheet.cell(row = 1, column = i).value + ": ")
            sheet.cell(row = line, column = i).value = float(add)
    directory = "C:\\Users\\shrey\\Desktop"
    directory = directory.lower()
    os.chdir(directory)
    workbook.save("Stocks.xlsx")

def main():
    answer = 'Y'
    while answer == 'Y':
        setup()
        answer = input("Continue? (Y/N)  ")

main()