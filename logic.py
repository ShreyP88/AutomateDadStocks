# Process:
# Import necessary libraries
# Get spreadsheet location and name
# Store values in variables
# Make the input text of the name of the stock uppercase 
# Present a small menu which will allow to either edit, add, or view
# Create functions for each with proper logic

import openpyxl
import os

from openpyxl import workbook
from openpyxl import load_workbook

def setup():
    directory = "C:\\Users\\shrey\\Desktop"
    directory = directory.lower()
    os.chdir(directory)
    spreadname = "Stocks.xlsx"
    workbook = openpyxl.load_workbook(spreadname)
    """ sheet_Arr = workbook.sheetnames
    answer = ''
    sheetname = ''
    while answer != 'Y':
        for x in sheet_Arr:
            answer = input(x + ": Is this the sheetname which you are trying to access? (Y/N)  ")
            answer = answer.upper()
            if (answer == 'Y'):
                sheetname = sheet_Arr[0]
                break
    print(sheetname)
     choose(sheet, workbook) """
    sheet = workbook["Sheet1"]
    return sheet
   

def showStock(stockName):
    sheet = setup()
    stockName = stockName.upper()
    stockRow = 0
    toReturn = ''
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column = 2).value == stockName:
            stockRow = i
    toReturn += '\n'
    toReturn += "Here is the info for " + str(stockName) + ":" + '\n'
    for i in range(1, sheet.max_column + 1):
        if i == 8 or i == 3 or i == 2 or i == 1:
            toReturn += str(i) + ": " + sheet.cell(row = 1, column = i).value + ": " + str(sheet.cell(row = stockRow, column = i).value)
        else:
             toReturn += str(i) + ": " + sheet.cell(row = 1, column = i).value + ": " + str(float((sheet.cell(row = stockRow, column = i).value)))
        toReturn += '\n'
    return toReturn

def showCurrentStocks() -> str:
    sheet = setup()
    toReturn = ''
    toReturn += "Here is a current list of your stocks: " + '\n'
    for i in range(2, sheet.max_row + 1):
        toReturn += "Investment Name: " + str(sheet.cell(row = i, column = 1).value) + '\n' + "Stock Symbol: " + str(sheet.cell(row = i, column = 2).value)
        toReturn += '\n'
    return toReturn
    
def getStockRow(stock):
    sheet = setup()
    stockRow = 1
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column = 2).value == stock:
            stockRow = i
    return stockRow

def editStock(choice, stockSymbol, edit):
    directory = "C:\\Users\\shrey\\Desktop"
    directory = directory.lower()
    os.chdir(directory)
    spreadname = "Stocks.xlsx"
    workbook = openpyxl.load_workbook(spreadname)
    sheet = workbook["Sheet1"]
    stockRow = getStockRow(stockSymbol)

    if choice == 7 or choice == 2 or choice == 3 or choice == 1:
        sheet.cell(row = stockRow, column = choice).value = edit
    else:
        sheet.cell(row = stockRow, column = choice).value = float(edit)
    workbook.save("Stocks.xlsx")

def addStock(sheet, workbook):
    line = sheet.max_row + 1
    for i in range(1, sheet.max_column + 1):
        if i == 7 or i == 2 or i == 1:
            add = input(sheet.cell(row = 1, column = i).value + ": ")
            sheet.cell(row = line, column = i).value = add
        else:
            add = input(sheet.cell(row = 1, column = i).value + ": ")
            sheet.cell(row = line, column = i).value = float(add)
    directory = "C:\\Users\\shrey\\Desktop"
    directory = directory.lower()
    os.chdir(directory)
    workbook.save("Stocks.xlsx")

def main():
    answer = 'Y'
    while answer == 'Y':
        setup()
        answer = input("Continue? (Y/N)  ")

if __name__ == "__main__":
    main()
